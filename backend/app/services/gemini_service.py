import google.generativeai as genai
from typing import Dict, Any, List
from app.config import settings
import json
import io
import time
import asyncio
import traceback
from PyPDF2 import PdfReader
from openpyxl import load_workbook

class GeminiService:
    """Service for AI analysis using Google Gemini"""
    
    def __init__(self):
        genai.configure(api_key=settings.GEMINI_API_KEY)
        self.model = genai.GenerativeModel(settings.GEMINI_MODEL)

    def _get_field_descriptions_led(self, criterion_num: int) -> str:
        """Get detailed field descriptions for LED data"""
        descriptions = {
            1: '''- vmts_unik_spesifik (string): Jelaskan secara singkat keunikan/kekhasan VMTS berdasarkan teks.
- vmts_dukungan_renstra_kurikulum (string): Jelaskan secara singkat bagaimana VMTS didukung oleh renstra dan kurikulum.
- vmts_stakeholder_internal (string): Sebutkan bukti keterlibatan stakeholder internal (dosen, mahasiswa, tendik) dalam penyusunan VMTS.
- vmts_stakeholder_eksternal (string): Sebutkan bukti keterlibatan stakeholder eksternal (lulusan, pengguna) dalam penyusunan VMTS.
- vmts_sosialisasi (string): Jelaskan secara singkat metode sosialisasi VMTS yang dijelaskan dalam dokumen.
- vmts_pemahaman (string): Jelaskan bukti adanya pemahaman VMTS oleh sivitas akademika.
- vmts_pencapaian_konkret (string): Sebutkan satu atau dua contoh pencapaian konkret dari implementasi VMTS.
- vmts_dampak_berkelanjutan (string): Jelaskan dampak berkelanjutan yang dihasilkan dari pencapaian VMTS.''',
            2: '''- tata_pamong_kelengkapan (string): Jelaskan bukti kelengkapan struktur organisasi dan tata pamong.
- tata_pamong_governance (string): Jelaskan bukti penerapan prinsip Good University Governance (transparan, akuntabel, dll).
- komitmen_pimpinan (string): Jelaskan bukti komitmen pimpinan UPPS.
- kemampuan_manajerial (string): Jelaskan bukti kemampuan manajerial pimpinan UPPS.
- pengelolaan_keuangan (string): Jelaskan bukti adanya sistem pengelolaan keuangan yang transparan dan akuntabel.''',
            3: '''- pemutakhiran_kurikulum (string): Jelaskan bukti adanya proses pemutakhiran kurikulum yang melibatkan stakeholder.
- kesesuaian_profil_cpl (string): Jelaskan kesesuaian antara profil lulusan dengan CPL.
- rps_kelengkapan (string): Jelaskan kelengkapan komponen dalam dokumen RPS.
- proses_pembelajaran_efektivitas (string): Jelaskan metode pembelajaran yang berpusat pada mahasiswa (SCL).
- suasana_akademik_pengelolaan (string): Jelaskan cara UPPS mengelola dan mengembangkan suasana akademik.
- kesesuaian_penelitian (string): Jelaskan kesesuaian penelitian dengan peta jalan.
- kesesuaian_pkm (string): Jelaskan kesesuaian PkM dengan peta jalan.''',
            5: '''- sarana_prasarana_akademik (string): Deskripsikan secara singkat sarana prasarana utama untuk kegiatan akademik.
- sarana_prasarana_non_akademik (string): Deskripsikan secara singkat sarana prasarana untuk kegiatan non-akademik.
- k3l (string): Jelaskan bukti implementasi sistem K3L.''',
            7: '''- keberadaan_unit_spmi (string): Sebutkan nama unit penjaminan mutu yang dijelaskan.
- ketersediaan_perangkat_spmi (string): Sebutkan dokumen-dokumen SPMI yang tersedia.
- keterlaksanaan_spmi (string): Jelaskan bukti keterlaksanaan siklus SPMI (PPEPP).
- evaluasi_capaian_kinerja (string): Jelaskan mekanisme evaluasi capaian kinerja.
- kepuasan_pemangku_kepentingan (string): Jelaskan metode pengukuran kepuasan pemangku kepentingan.'''
        }
        return descriptions.get(criterion_num, "No LED fields for this criterion")
"No LED fields for this criterion")

    def _get_lkps_extraction_prompt(self, criterion_num: int, lkps_content_snippet: str) -> str:
        """Generates a highly specific and structured prompt for LKPS data extraction for a given criterion."""
        
        prompts = {
            2: {
                "fields": {
                    "kerjasama_pendidikan": "(number) Jumlah kerjasama PENDIDIKAN dari Tabel 4. Sum dari TS, TS-1, TS-2.",
                    "kerjasama_penelitian": "(number) Jumlah kerjasama PENELITIAN dari Tabel 4. Sum dari TS, TS-1, TS-2.",
                    "kerjasama_pkm": "(number) Jumlah kerjasama PKM dari Tabel 4. Sum dari TS, TS-1, TS-2.",
                    "kerjasama_internasional": "(number) Jumlah kerjasama tingkat INTERNASIONAL (RI) dari Tabel 4. Sum dari TS, TS-1, TS-2.",
                    "kerjasama_nasional": "(number) Jumlah kerjasama tingkat NASIONAL (RN) dari Tabel 4. Sum dari TS, TS-1, TS-2.",
                    "kerjasama_wilayah": "(number) Jumlah kerjasama tingkat LOKAL/WILAYAH (RW) dari Tabel 4. Sum dari TS, TS-1, TS-2.",
                },
                "example": {"kerjasama_pendidikan": 131, "kerjasama_internasional": 34}
            },
            3: {
                "fields": {
                    "persentase_bahan_ajar_penelitian_pkm": "(number) Persentase bahan ajar dari penelitian/PkM (0-100) dari Butir 14. Cari angka persentase.",
                    "pjp": "(number) PJP - Persentase pembelajaran berbasis praktik (0-100) dari Butir 17. Cari angka persentase.",
                    "basic_sciences_sks": "(number) Jumlah SKS mata kuliah sains dasar dari Butir 18. Cari angka total SKS.",
                    "ppdmhs": "(number) PPDMHS - Persentase praktik dalam mata kuliah sains (0-100) dari Butir 19. Cari angka persentase.",
                    "pkdmhs": "(number) PKDMHS - Persentase kerja praktik/magang (0-100) dari Butir 20. Cari angka persentase."
                },
                "example": {"persentase_bahan_ajar_penelitian_pkm": 20.0, "pjp": 30, "basic_sciences_sks": 24, "ppdmhs": 15, "pkdmhs": 10}
            },
            4: {
                "fields": {
                    "ndtps": "(number) NDTPS - Jumlah Dosen Tetap Program Studi. Cari di 'Tabel 3.a.1' pada baris 'Jumlah' atau 'Total'.",
                    "pdtt": "(number) PDTT - Persentase dosen tidak tetap (0-100). Cari di 'Butir 24' atau hitung dari 'Tabel 3.a.1' jika ada data dosen tidak tetap.",
                    "pds3": "(number) PDS3 - Persentase dosen S3/doktor (0-100). Cari di 'Tabel 3.a.1' pada kolom 'Pendidikan' atau 'Kualifikasi Akademik', hitung persentase S3 dari total dosen.",
                    "pgblkl": "(number) PGBLKL - Persentase Guru Besar + Lektor Kepala (0-100). Cari di 'Tabel 3.a.1' pada kolom 'Jabatan Akademik', hitung persentase GB+LK dari total dosen.",
                    "rbk_dtps": "(number) RBK - Rata-rata beban kerja DTPS per semester dari Butir 27. Cari angka rata-rata.",
                    "kinerja_penelitian_dtps_ri": "(number) Jumlah penelitian DTPS tingkat INTERNASIONAL (RI) dari Tabel 3.b.1. Sum dari TS, TS-1, TS-2.",
                    "kinerja_penelitian_dtps_rn": "(number) Jumlah penelitian DTPS tingkat NASIONAL (RN) dari Tabel 3.b.1. Sum dari TS, TS-1, TS-2.",
                    "kinerja_penelitian_dtps_rw": "(number) Jumlah penelitian DTPS tingkat LOKAL (RW) dari Tabel 3.b.1. Sum dari TS, TS-1, TS-2.",
                    "kinerja_pkm_dtps_ri": "(number) Jumlah PkM DTPS tingkat INTERNASIONAL (RI) dari Tabel 3.b.1. Sum dari TS, TS-1, TS-2.",
                    "kinerja_pkm_dtps_rn": "(number) Jumlah PkM DTPS tingkat NASIONAL (RN) dari Tabel 3.b.1. Sum dari TS, TS-1, TS-2.",
                    "kinerja_pkm_dtps_rw": "(number) Jumlah PkM DTPS tingkat LOKAL (RW) dari Tabel 3.b.1. Sum dari TS, TS-1, TS-2.",
                    "publikasi_ilmiah_dtps_ri": "(number) Jumlah publikasi ilmiah DTPS tingkat INTERNASIONAL (RI) dari Tabel 3.b.2. Sum dari TS, TS-1, TS-2.",
                    "publikasi_ilmiah_dtps_rn": "(number) Jumlah publikasi ilmiah DTPS tingkat NASIONAL (RN) dari Tabel 3.b.2. Sum dari TS, TS-1, TS-2.",
                    "publikasi_ilmiah_dtps_rw": "(number) Jumlah publikasi ilmiah DTPS tingkat LOKAL (RW) dari Tabel 3.b.2. Sum dari TS, TS-1, TS-2.",
                    "rlp_dtps": "(number) RLP - Rasio luaran penelitian per DTPS dari Butir 33. Cari angka rasio."
                },
                "example": {"ndtps": 29, "pdtt": 5, "pds3": 100.0, "pgblkl": 82.76, "kinerja_penelitian_dtps_rn": 12, "rlp_dtps": 0.5}
            },
            6: {
                "fields": {
                    "rmd": "(number) RMD - Rasio jumlah mahasiswa terhadap NDTPS dari Butir 37. Cari angka rasio.",
                    "pma": "(number) PMA - Persentase mahasiswa asing (0-100) dari Butir 38. Cari angka persentase.",
                    "ripk": "(number) RIPK - Rata-rata IPK lulusan (0.0-4.0) dari Butir 39. Cari angka rata-rata.",
                    "prestasi_akademik_ri": "(number) Jumlah prestasi AKADEMIK tingkat INTERNASIONAL (RI) dari Tabel 6.a. Sum dari TS, TS-1, TS-2.",
                    "prestasi_akademik_rn": "(number) Jumlah prestasi AKADEMIK tingkat NASIONAL (RN) dari Tabel 6.a. Sum dari TS, TS-1, TS-2.",
                    "prestasi_akademik_rw": "(number) Jumlah prestasi AKADEMIK tingkat LOKAL (RW) dari Tabel 6.a. Sum dari TS, TS-1, TS-2.",
                    "prestasi_non_akademik_ri": "(number) Jumlah prestasi NON-AKADEMIK tingkat INTERNASIONAL (RI) dari Tabel 6.a. Sum dari TS, TS-1, TS-2.",
                    "prestasi_non_akademik_rn": "(number) Jumlah prestasi NON-AKADEMIK tingkat NASIONAL (RN) dari Tabel 6.a. Sum dari TS, TS-1, TS-2.",
                    "prestasi_non_akademik_rw": "(number) Jumlah prestasi NON-AKADEMIK tingkat LOKAL (RW) dari Tabel 6.a. Sum dari TS, TS-1, TS-2.",
                    "ptw": "(number) PTW - Persentase lulusan tepat waktu (0-100) dari Butir 47. Cari angka persentase.",
                    "publikasi_mahasiswa_ri": "(number) Jumlah publikasi mahasiswa INTERNASIONAL (RI) dari Tabel 6.b. Sum dari TS, TS-1, TS-2.",
                    "publikasi_mahasiswa_rn": "(number) Jumlah publikasi mahasiswa NASIONAL (RN) dari Tabel 6.b. Sum dari TS, TS-1, TS-2.",
                    "publikasi_mahasiswa_rw": "(number) Jumlah publikasi mahasiswa LOKAL (RW) dari Tabel 6.b. Sum dari TS, TS-1, TS-2.",
                    "wt": "(number) WT - Rata-rata waktu tunggu kerja dalam bulan dari Butir 50. Cari angka rata-rata.",
                    "kbk": "(number) KBK - Persentase kesesuaian bidang kerja (0-100) dari Butir 51. Cari angka persentase.",
                    "tingkat_tempat_kerja_ri": "(number) Jumlah lulusan bekerja tingkat INTERNASIONAL dari Butir 51. Sum dari TS, TS-1, TS-2.",
                    "tingkat_tempat_kerja_rn": "(number) Jumlah lulusan bekerja tingkat NASIONAL dari Butir 51. Sum dari TS, TS-1, TS-2.",
                    "tingkat_tempat_kerja_rw": "(number) Jumlah lulusan bekerja tingkat LOKAL dari Butir 51. Sum dari TS, TS-1, TS-2.",
                    "kepuasan_pengguna_a1": "(number) Jumlah responden SANGAT BAIK untuk aspek 1 (Etika) dari Tabel Kepuasan Pengguna. Sum dari TS, TS-1, TS-2.",
                    "kepuasan_pengguna_b1": "(number) Jumlah responden BAIK untuk aspek 1 (Etika). Sum dari TS, TS-1, TS-2.",
                    "kepuasan_pengguna_c1": "(number) Jumlah responden CUKUP untuk aspek 1 (Etika). Sum dari TS, TS-1, TS-2.",
                    "kepuasan_pengguna_d1": "(number) Jumlah responden KURANG untuk aspek 1 (Etika). Sum dari TS, TS-1, TS-2."
                },
                "example": {"rmd": 20, "pma": 5, "ripk": 3.51, "prestasi_akademik_rn": 3, "wt": 3, "kbk": 75, "tingkat_tempat_kerja_rn": 50, "kepuasan_pengguna_a1": 50}
            }
        }

        prompt_data = prompts.get(criterion_num)
        if not prompt_data:
            return None

        field_definitions = "\n".join([f'- `{key}`: {desc}' for key, desc in prompt_data["fields"].items()])

        prompt = f"""# OBJECTIVE
Extract specific NUMERICAL data points from the provided LKPS document snippet for Criterion {criterion_num}.

# EXTRACTION RULES
1. Analyze the text, focusing on tables and numbered lists (`Butir`).
2. For each field, find the corresponding numerical value. If data spans multiple years (TS, TS-1, TS-2), SUM the values unless it's a ratio or percentage.
3. **RETURN ONLY NUMBERS.** Do not return text, descriptions, or formulas.
4. If a value is not found or not applicable, use `0`.
5. Ensure percentages are numbers from 0-100 (e.g., "85%" -> `85.0`).
6. For fields like PDS3, PGBLKL, calculate the percentage from the raw counts in the table.

# LKPS DOCUMENT SNIPPET
```
{lkps_content_snippet}
```

# FIELDS TO EXTRACT (Criterion {criterion_num})
{field_definitions}

# OUTPUT FORMAT
Return **ONLY a single JSON object** with the key `lkps_data` containing the extracted fields. Do not add any explanations, markdown, or other text.

Example:
```json
{{
  "lkps_data": {json.dumps(prompt_data["example"])}
}}
```
"""
        return prompt

    def _get_specific_lkps_prompt(self, lkps_content_snippet: str, fields: Dict[str, str], example: Dict[str, Any]) -> str:
        """Generates a focused LKPS extraction prompt for a specific subset of fields."""
        field_definitions = "\n".join([f'- `{key}`: {desc}' for key, desc in fields.items()])

        prompt = f"""# OBJECTIVE
Extract specific NUMERICAL data points from the provided LKPS document snippet.

# EXTRACTION RULES
1. Analyze the text, focusing on tables (especially "Tabel 4").
2. For each field, find the corresponding numerical value. SUM values across years (TS, TS-1, TS-2) if present.
3. **RETURN ONLY NUMBERS.** Do not return text.
4. If a value is not found, use `0`.

# LKPS DOCUMENT SNIPPET
```
{lkps_content_snippet}
```

# FIELDS TO EXTRACT
{field_definitions}

# OUTPUT FORMAT
Return **ONLY a single JSON object** with the key `lkps_data` containing the extracted fields.

Example:
```json
{{
  "lkps_data": {json.dumps(example)}
}}
```
"""
        return prompt

    async def verify_document_type(
        self,
        filename: str,
        file_content: bytes,
        expected_type: str
    ) -> Dict[str, Any]:
        """
        Verify if uploaded document matches expected type (LED/LKPS)
        """
        try:
            text_content = ""
            if filename.lower().endswith('.pdf'):
                text_content = await self.extract_text_from_pdf(file_content)
            elif filename.lower().endswith(('.xlsx', '.xls', '.csv')):
                text_content = await self.extract_text_from_excel(file_content, filename)
            
            text_lower = text_content.lower()
            
            if expected_type == "LED":
                led_keywords = ["laporan evaluasi diri", "led", "evaluasi diri", "program studi", "akreditasi", "visi misi"]
                if any(keyword in text_lower for keyword in led_keywords):
                    return {"isValid": True, "confidence": 0.95, "detectedType": "LED", "reason": "Dokumen mengandung format LED."}
            
            elif expected_type == "LKPS":
                lkps_keywords = ["laporan kinerja program studi", "lkps", "lam-tek", "data dosen", "kinerja program"]
                if any(keyword in text_lower for keyword in lkps_keywords):
                    return {"isValid": True, "confidence": 0.95, "detectedType": "LKPS", "reason": "Dokumen mengandung format LKPS."}

            return {"isValid": False, "confidence": 0.4, "detectedType": "OTHER", "reason": f"Dokumen tidak mengandung kata kunci untuk {expected_type}."}

        except Exception as e:
            print(f"Error in document verification: {e}")
            return {"isValid": False, "confidence": 0.0, "detectedType": "ERROR", "reason": str(e)}

    async def analyze_documents_for_scoring(
        self,
        program_studi: str,
        institusi: str,
        led_content: str,
        lkps_content: str,
        program_type: str = "S"
    ) -> Dict[str, Any]:
        """
        LAM-TEK 2025 Accreditation Assessor AI for 7 criteria.
        """
        print(f"[LAM-TEK] Starting LAM-TEK 2025 analysis for {program_studi} ({program_type}) - {institusi}")

        criteria_config = {
            1: {"name": "Diferensiasi Misi", "led_keys": ["vmts_unik_spesifik", "vmts_dukungan_renstra_kurikulum", "vmts_linearitas_pt", "vmts_kesesuaian_renstra", "vmts_kesesuaian_kurikulum", "vmts_stakeholder_internal", "vmts_stakeholder_eksternal", "vmts_sosialisasi", "vmts_pemahaman", "vmts_pencapaian_konkret", "vmts_dampak_berkelanjutan"], "lkps_keys": []},
            2: {"name": "Akuntabilitas", "led_keys": ["tata_pamong_kelengkapan", "tata_pamong_governance", "komitmen_pimpinan", "kemampuan_manajerial", "pengelolaan_keuangan"], "lkps_keys": ["jumlah_dtps", "kerjasama_pendidikan", "kerjasama_penelitian", "kerjasama_pkm", "kerjasama_internasional", "kerjasama_nasional", "kerjasama_wilayah"]},
            3: {"name": "Relevansi Pendidikan, Penelitian, dan PkM", "led_keys": ["pemutakhiran_kurikulum", "profil_lulusan", "kesesuaian_profil_cpl", "kesesuaian_cpl_standar_kompetensi", "rps_kelengkapan", "rps_tinjauan_rutin", "proses_pembelajaran_efektivitas", "proses_pembelajaran_tinjauan_rutin", "capstone_design", "suasana_akademik_pengelolaan", "suasana_akademik_integritas", "kesesuaian_penelitian", "kesesuaian_pkm"], "lkps_keys": ["persentase_bahan_ajar_penelitian_pkm", "pjp", "basic_sciences_sks", "ppdmhs", "pkdmhs"]},
            4: {"name": "Sumber Daya Manusia", "led_keys": [], "lkps_keys": ["ndtps", "pdtt", "pds3", "pgblkl", "rbk_dtps", "kinerja_penelitian_dtps_ri", "kinerja_penelitian_dtps_rn", "kinerja_penelitian_dtps_rw", "kinerja_pkm_dtps_ri", "kinerja_pkm_dtps_rn", "kinerja_pkm_dtps_rw", "publikasi_ilmiah_dtps_ri", "publikasi_ilmiah_dtps_rn", "publikasi_ilmiah_dtps_rw", "rlp_dtps"]},
            5: {"name": "Sarana, Prasarana, dan K3L", "led_keys": ["sarana_prasarana_akademik", "sarana_prasarana_non_akademik", "k3l"], "lkps_keys": []},
            6: {"name": "Mahasiswa dan Luaran Mahasiswa", "led_keys": [], "lkps_keys": ["rmd", "pma", "ripk", "prestasi_akademik_ri", "prestasi_akademik_rn", "prestasi_akademik_rw", "prestasi_non_akademik_ri", "prestasi_non_akademik_rn", "prestasi_non_akademik_rw", "ptw", "publikasi_mahasiswa_ri", "publikasi_mahasiswa_rn", "publikasi_mahasiswa_rw", "wt", "kbk", "tingkat_tempat_kerja_ri", "tingkat_tempat_kerja_rn", "tingkat_tempat_kerja_rw"] + [f'kepuasan_pengguna_{cat}{i}' for cat in "abcd" for i in range(1, 2)]},
            7: {"name": "Sistem Penjaminan Mutu", "led_keys": ["keberadaan_unit_spmi", "ketersediaan_perangkat_spmi", "keterlaksanaan_spmi", "evaluasi_capaian_kinerja", "kepuasan_pemangku_kepentingan"], "lkps_keys": []}
        }

        return await self._multi_request_extraction(
            led_content, lkps_content, program_studi, institusi, program_type, criteria_config
        )

    def _find_relevant_snippet(self, content: str, keywords: List[str], window_size: int = 15000) -> str:
        """
        Finds snippets of text around given keywords.
        """
        snippets = []
        found_keywords = set()
        lower_content = content.lower()

        for keyword in keywords:
            try:
                idx = lower_content.find(keyword.lower())
                
                if idx != -1:
                    is_new = True
                    for found_key in found_keywords:
                        found_key_idx = lower_content.find(found_key.lower())
                        if abs(found_key_idx - idx) < window_size / 2:
                            is_new = False
                            break
                    
                    if is_new:
                        start = max(0, idx - int(window_size * 0.25))
                        end = min(len(content), idx + int(window_size * 0.75))
                        snippets.append(f"--- SNIPPET RELEVAN DARI DOKUMEN UNTUK KATA KUNCI '{keyword}' ---\n{content[start:end]}")
                        found_keywords.add(keyword)

            except Exception:
                continue

        if snippets:
            print(f"[LAM-TEK] Ditemukan snippet relevan untuk kata kunci: {list(found_keywords)}")
            return "\n\n".join(snippets)
        else:
            print(f"[LAM-TEK] Peringatan: Tidak ada kata kunci relevan yang ditemukan. Menggunakan {window_size} karakter pertama.")
            return content[:window_size]

    async def _multi_request_extraction(
        self, led_content, lkps_content, program_studi, institusi, program_type, criteria_config
    ) -> Dict[str, Any]:
        final_led_data = {}
        final_lkps_data = {}
        errors = []

        print(f"[LAM-TEK] Memulai ekstraksi multi-permintaan dengan alur baru...")

        # --- LANGKAH 1: Ekstraksi Kriteria 4 (SDM) untuk mendapatkan NDTPS ---
        print("[LAM-TEK] [PRIORITAS] Menganalisis Kriteria 4: Sumber Daya Manusia untuk NDTPS...")
        try:
            lkps_keyword_map_k4 = ["Tabel 3.a.1", "Dosen Tetap", "NDTPS"]
            lkps_snippet_k4 = self._find_relevant_snippet(lkps_content, lkps_keyword_map_k4, window_size=20000)
            lkps_prompt_k4 = self._get_lkps_extraction_prompt(4, lkps_snippet_k4)
            
            if lkps_prompt_k4:
                lkps_response_text_k4 = await self._generate_gemini_response(lkps_prompt_k4)
                lkps_data_k4 = self._parse_json_response(lkps_response_text_k4, "lkps_data")
                final_lkps_data.update(lkps_data_k4)
                ndtps_value = final_lkps_data.get("ndtps", 0)
                print(f"[LAM-TEK] ✓ Ekstraksi NDTPS selesai. Nilai NDTPS: {ndtps_value}")
            else:
                errors.append("Gagal membuat prompt untuk Kriteria 4.")

        except Exception as e:
            error_msg = f"Kriteria 4 (Prioritas NDTPS) gagal: {str(e)}"
            print(f"[LAM-TEK] ✗ {error_msg}")
            errors.append(error_msg)

        # --- LANGKAH 2: Loop untuk kriteria lainnya ---
        lkps_keyword_map = {
            2: ["Tabel 4", "Kerjasama", "Jumlah Kerjasama", "Mitra Kerjasama"],
            3: ["Butir 14", "Butir 17", "integrasi penelitian", "praktikum"],
            6: ["Tabel 6.a", "Tabel 6.b", "Prestasi Akademik", "Publikasi Mahasiswa", "Kepuasan Pengguna"],
        }

        criteria_to_process = [1, 2, 3, 5, 6, 7] # Semua kecuali 4
        for i in criteria_to_process:
            criterion = criteria_config[i]
            criterion_name = criterion["name"]
            
            await asyncio.sleep(2) # Jeda antar permintaan

            # Special handling for Criterion 2 to avoid MAX_TOKENS error
            if i == 2:
                print("[LAM-TEK] Menerapkan strategi permintaan terpisah untuk Kriteria 2...")
                try:
                    # Define the snippet first, as it's needed for both sub-requests
                    keywords = lkps_keyword_map.get(i, [])
                    lkps_snippet = self._find_relevant_snippet(lkps_content, keywords, window_size=15000)

                    # Request 2a: Pendidikan, Penelitian, PkM
                    fields_2a = {"kerjasama_pendidikan": "(number) Jumlah kerjasama PENDIDIKAN.", "kerjasama_penelitian": "(number) Jumlah kerjasama PENELITIAN.", "kerjasama_pkm": "(number) Jumlah kerjasama PKM."}
                    prompt_2a = self._get_specific_lkps_prompt(lkps_snippet, fields_2a, {"kerjasama_pendidikan": 10, "kerjasama_penelitian": 5})
                    response_2a = await self._generate_gemini_response(prompt_2a)
                    data_2a = self._parse_json_response(response_2a, "lkps_data")
                    final_lkps_data.update(data_2a)
                    print("[LAM-TEK] ✓ Bagian 2a (Pendidikan, Penelitian, PkM) selesai.")

                    await asyncio.sleep(2) # Jeda kecil

                    # Request 2b: Internasional, Nasional, Wilayah
                    fields_2b = {"kerjasama_internasional": "(number) Jumlah kerjasama tingkat INTERNASIONAL (RI).", "kerjasama_nasional": "(number) Jumlah kerjasama tingkat NASIONAL (RN).", "kerjasama_wilayah": "(number) Jumlah kerjasama tingkat LOKAL/WILAYAH (RW)."}
                    prompt_2b = self._get_specific_lkps_prompt(lkps_snippet, fields_2b, {"kerjasama_internasional": 4, "kerjasama_nasional": 20})
                    response_2b = await self._generate_gemini_response(prompt_2b)
                    data_2b = self._parse_json_response(response_2b, "lkps_data")
                    final_lkps_data.update(data_2b)
                    print("[LAM-TEK] ✓ Bagian 2b (Internasional, Nasional, Wilayah) selesai.")

                    # Handle LED data for Kriteria 2 as well
                    if criterion["led_keys"]:
                        led_snippet = led_content[:25000]
                        led_prompt = self._get_led_extraction_prompt(i, led_snippet)
                        led_response_text = await self._generate_gemini_response(led_prompt)
                        led_data = self._parse_json_response(led_response_text, "led_data")
                        final_led_data.update(led_data)

                    print(f"[LAM-TEK] ✓ Ekstraksi Kriteria {i} selesai.")
                    continue # Lanjutkan ke loop berikutnya
                except Exception as e:
                    error_msg = f"Kriteria {i} ({criterion_name}) gagal: {str(e)}"
                    print(f"[LAM-TEK] ✗ {error_msg}")
                    errors.append(error_msg)
                    continue

        print(f"[LAM-TEK] Ekstraksi multi-permintaan selesai! (7 Kriteria LAM-TEK 2025)")
        print(f"[LAM-TEK] - Ekstrak {len(final_led_data)} field LED")
        print(f"[LAM-TEK] - Ekstrak {len(final_lkps_data)} field LKPS")

        return {
            "led_data": final_led_data,
            "lkps_data": final_lkps_data,
            "scoring_readiness": {"ready_for_lamtek_scoring": not errors, "error": "; ".join(errors) if errors else None}
        }

    def _get_led_extraction_prompt(self, criterion_num: int, led_content_snippet: str) -> str:
        field_definitions = self._get_field_descriptions_led(criterion_num)
        return f"""# OBJECTIVE
Extract qualitative information from the LED document for Criterion {criterion_num}.

# LED DOCUMENT SNIPPET
```
{led_content_snippet}
```

# FIELDS TO EXTRACT
Based on the text, provide a summary or boolean value for each field:
{field_definitions}

# OUTPUT FORMAT
Return **ONLY a single JSON object** with the key `led_data`.
Example:
```json
{{
  "led_data": {{
    "vmts_unik_spesifik": "Visi PS S2 TIP adalah menjadi program studi unggul...",
    "vmts_sosialisasi": true
  }}
}}
```
"""

    async def _generate_gemini_response(self, prompt: str) -> str:
        try:
            generation_config = {
                "temperature": 0.0,
                "max_output_tokens": 8192,
            }
            response = await self.model.generate_content_async(prompt, generation_config=generation_config)
            
            # Robustness check: Handle empty responses from the API
            if not response.parts:
                print(f"[LAM-TEK] ⚠️ Peringatan: Menerima respons kosong dari API.")
                try:
                    # Log the reason if available, often due to safety filters
                    finish_reason = response.candidates[0].finish_reason
                    print(f"[LAM-TEK] Alasan Selesai (Finish Reason): {finish_reason.name}")
                except (IndexError, AttributeError):
                    pass # Ignore if we can't get the reason
                return '{}' # Return empty JSON object to prevent crash

            return response.text
        except Exception as e:
            print(f"[LAM-TEK] ✗ Error saat menghasilkan konten Gemini: {e}")
            # In case of other errors, also return an empty JSON to be safe
            return '{}'

    def _parse_json_response(self, response_text: str, data_key: str) -> dict:
        try:
            start = response_text.find('{')
            end = response_text.rfind('}')
            if start == -1 or end == -1 or end < start:
                print(f"[LAM-TEK] ✗ Could not find a valid JSON object in the response for key '{data_key}'.")
                print(f"[LAM-TEK] Problematic response snippet: {response_text[:500]}")
                return {}

            json_str = response_text[start:end+1]
            data = json.loads(json_str)
            
            if data_key in data:
                return data.get(data_key, {})
            else:
                return data

        except (json.JSONDecodeError, IndexError) as e:
            print(f"[LAM-TEK] ✗ JSON parsing failed for key '{data_key}': {e}")
            print(f"[LAM-TEK] Problematic response snippet: {response_text[:500]}")
            return {}

    async def extract_text_from_pdf(self, pdf_content: bytes) -> str:
        text = ""
        try:
            pdf_file = io.BytesIO(pdf_content)
            reader = PdfReader(pdf_file)
            for page in reader.pages:
                text += page.extract_text() or ""
        except Exception as e:
            print(f"Error extracting text from PDF: {e}")
        return text
    
    async def extract_text_from_excel(self, file_content: bytes, filename: str) -> str:
        text = ""
        try:
            file_io = io.BytesIO(file_content)
            
            if filename.lower().endswith('.csv'):
                # Handle CSV files
                file_io.seek(0)
                # Decode bytes to string for csv.reader
                csv_text = file_io.read().decode('utf-8', errors='replace')
                reader = csv.reader(csv_text.splitlines())
                for row in reader:
                    # Join cells with a space, and rows with a newline
                    text += " ".join(cell for cell in row if cell) + "\n"
                print(f"[Gemini] DEBUG: First 1000 chars of CSV extracted text:\n{text[:1000]}")
            
            elif filename.lower().endswith(('.xlsx', '.xls')):
                # Handle Excel files with openpyxl
                workbook = load_workbook(file_io)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        row_text = []
                        for cell in row:
                            if cell.value:
                                row_text.append(str(cell.value))
                        if row_text:
                            text += " ".join(row_text) + "\n"
            else:
                print(f"[Gemini] Unsupported file type for excel extraction: {filename}")

        except Exception as e:
            print(f"Error extracting text from Excel/CSV: {e}")
            traceback.print_exc()
        return text

gemini_service = GeminiService()